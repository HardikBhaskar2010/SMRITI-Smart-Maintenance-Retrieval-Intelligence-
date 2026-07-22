"""
Phase 2 Export Engine — PDF and Excel report generation.
"""
import io
import logging
from datetime import datetime, timezone

logger = logging.getLogger(__name__)


def _severity_color_rgb(severity: str) -> tuple:
    return {
        "CRITICAL": (0.94, 0.27, 0.27),
        "WARNING":  (0.96, 0.62, 0.04),
        "OK":       (0.13, 0.77, 0.37),
    }.get(severity, (0.5, 0.5, 0.5))


def generate_asset_pdf(asset_id: str, asset_data: dict, thread_items: list[dict]) -> bytes:
    """
    Generate a PDF Knowledge Debt Report for a single asset.
    Returns raw PDF bytes.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            topMargin=20 * mm, bottomMargin=20 * mm,
            leftMargin=20 * mm, rightMargin=20 * mm,
        )
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle(
            "title",
            parent=styles["Title"],
            fontSize=22, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0D0F14"),
            spaceAfter=4,
        )
        heading_style = ParagraphStyle(
            "heading2",
            parent=styles["Heading2"],
            fontSize=14, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#13161E"),
            spaceAfter=6, spaceBefore=14,
        )
        normal = ParagraphStyle(
            "normal2",
            parent=styles["Normal"],
            fontSize=10, fontName="Helvetica",
            textColor=colors.HexColor("#2A2E3D"),
            spaceAfter=4,
        )

        # ── Header ──────────────────────────────────────────────────────
        story.append(Paragraph("SMRITI — Knowledge Debt Report", title_style))
        story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#00C9A7")))
        story.append(Spacer(1, 8 * mm))

        # Asset overview table
        debt_score = asset_data.get("debt_score", "N/A")
        severity = asset_data.get("severity", "UNKNOWN")
        r, g, b = _severity_color_rgb(severity)
        sev_color = colors.Color(r, g, b)

        overview_data = [
            ["Asset ID", asset_id],
            ["Asset Type", asset_data.get("asset_type", "N/A")],
            ["Knowledge Debt Score", f"{debt_score} / 100"],
            ["Risk Level", severity],
            ["Knowledge Items", str(asset_data.get("item_count", len(thread_items)))],
            ["Expert Count", str(asset_data.get("expert_count", "N/A"))],
            ["Last Updated", asset_data.get("last_updated", "N/A")[:10]],
            ["Report Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
        ]
        overview_table = Table(overview_data, colWidths=[60 * mm, 110 * mm])
        overview_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F4F5F7")),
            ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#E0E0E0")),
            ("PADDING",    (0, 0), (-1, -1), 6),
            ("TEXTCOLOR",  (1, 3), (1, 3), sev_color),
            ("FONTNAME",   (1, 3), (1, 3), "Helvetica-Bold"),
        ]))
        story.append(overview_table)
        story.append(Spacer(1, 6 * mm))

        # Score breakdown
        breakdown = asset_data.get("breakdown") or {}
        if breakdown:
            story.append(Paragraph("Score Breakdown", heading_style))
            bd_data = [["Dimension", "Penalty Points"]] + [
                [k.replace("_", " ").title(), str(v)]
                for k, v in breakdown.items()
            ]
            bd_table = Table(bd_data, colWidths=[100 * mm, 60 * mm])
            bd_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#00C9A7")),
                ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",   (0, 0), (-1, -1), 10),
                ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#E0E0E0")),
                ("PADDING",    (0, 0), (-1, -1), 6),
            ]))
            story.append(bd_table)
            story.append(Spacer(1, 6 * mm))

        # Knowledge Thread
        story.append(Paragraph("Knowledge Thread", heading_style))
        story.append(Paragraph(
            f"The following {len(thread_items)} knowledge items are stored for this asset:",
            normal,
        ))
        story.append(Spacer(1, 4 * mm))

        for i, item in enumerate(thread_items[:50], 1):  # Cap at 50 items
            source = item.get("source_document", "Unknown")
            page = item.get("source_page")
            section = item.get("source_section", "")
            content = item.get("content", "")[:400]
            attributed = "✓ Expert attributed" if item.get("expert_attributed") else "Auto-ingested"

            item_data = [
                [f"#{i}", f"{source} — p.{page}" if page else source],
                ["Section", section or "N/A"],
                ["Attribution", attributed],
                ["Content", Paragraph(content, normal)],
            ]
            item_table = Table(item_data, colWidths=[25 * mm, 145 * mm])
            item_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F0F9F7")),
                ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE",   (0, 0), (-1, -1), 9),
                ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor("#DDEEE9")),
                ("PADDING",    (0, 0), (-1, -1), 5),
                ("VALIGN",     (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(item_table)
            story.append(Spacer(1, 3 * mm))

        doc.build(story)
        return buf.getvalue()

    except ImportError:
        logger.warning("reportlab not installed — returning plain text PDF placeholder")
        content = f"SMRITI Knowledge Debt Report\nAsset: {asset_id}\nScore: {asset_data.get('debt_score')}\n"
        return content.encode()


def generate_asset_excel(asset_id: str, asset_data: dict, thread_items: list[dict]) -> bytes:
    """
    Generate an Excel workbook with asset summary + thread items.
    Returns raw .xlsx bytes.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Sheet 1: Summary ────────────────────────────────────────────
        ws_sum = wb.active
        ws_sum.title = "Asset Summary"

        teal = "FF00C9A7"
        dark = "FF0D0F14"
        header_font = Font(bold=True, color="FFFFFFFF", name="Calibri", size=11)
        teal_fill = PatternFill("solid", fgColor=teal)

        ws_sum.append(["SMRITI Knowledge Debt Report"])
        ws_sum["A1"].font = Font(bold=True, size=16, name="Calibri", color=dark)
        ws_sum.append(["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
        ws_sum.append([])

        summary_rows = [
            ("Asset ID", asset_id),
            ("Asset Type", asset_data.get("asset_type", "N/A")),
            ("Debt Score", asset_data.get("debt_score", "N/A")),
            ("Risk Level", asset_data.get("severity", "N/A")),
            ("Knowledge Items", asset_data.get("item_count", len(thread_items))),
            ("Expert Count", asset_data.get("expert_count", "N/A")),
            ("Last Updated", asset_data.get("last_updated", "N/A")[:10]),
        ]
        for row in summary_rows:
            ws_sum.append(list(row))
            ws_sum.cell(ws_sum.max_row, 1).font = Font(bold=True, name="Calibri")

        ws_sum.column_dimensions["A"].width = 22
        ws_sum.column_dimensions["B"].width = 35

        # ── Sheet 2: Knowledge Thread ───────────────────────────────────
        ws_items = wb.create_sheet("Knowledge Thread")
        headers = [
            "#", "Source Document", "Page", "Section",
            "Content", "Expert Attributed", "Expert Name", "Added At",
        ]
        ws_items.append(headers)
        for cell in ws_items[1]:
            cell.font = header_font
            cell.fill = teal_fill
            cell.alignment = Alignment(horizontal="center")

        for i, item in enumerate(thread_items, 1):
            ws_items.append([
                i,
                item.get("source_document", ""),
                item.get("source_page", ""),
                item.get("source_section", ""),
                item.get("content", "")[:500],
                "Yes" if item.get("expert_attributed") else "No",
                item.get("expert_name", ""),
                item.get("added_at", "")[:19],
            ])

        # Auto-width columns
        col_widths = [5, 30, 8, 20, 60, 18, 20, 20]
        for i, width in enumerate(col_widths, 1):
            ws_items.column_dimensions[get_column_letter(i)].width = width
        ws_items.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except ImportError:
        logger.warning("openpyxl not installed — returning CSV bytes")
        lines = ["Asset ID,Debt Score,Severity"]
        lines.append(f"{asset_id},{asset_data.get('debt_score')},{asset_data.get('severity')}")
        return "\n".join(lines).encode()
